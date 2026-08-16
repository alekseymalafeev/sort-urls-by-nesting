#!/usr/bin/env python3
# Sort URLs by Nesting by Malafeev Aleksey
# https://github.com/alekseymalafeev/sort-urls-by-nesting
# https://t.me/todaSE0

"""
Split a list of URLs into Excel sheets by URL nesting depth (path segments).
"""

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook

EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_DATA_ROWS = EXCEL_MAX_ROWS - 1  # reserve header row


def detect_csv_dialect(path: Path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        try:
            has_header = csv.Sniffer().has_header(sample)
        except csv.Error:
            has_header = False
    return dialect, has_header


def nesting_level(url: str) -> int:
    """
    Nesting depth by path segments.
    Examples:
      https://example.com/           -> Level 1
      https://example.com/catalog    -> Level 2
      https://example.com/a/b/c      -> Level 4
    """
    url = (url or "").strip()
    try:
        parsed = urlparse(url)
    except Exception:
        return 1
    segments = [s for s in (parsed.path or "/").split("/") if s]
    return max(1, 1 + len(segments))


class LevelWorkbook:
    def __init__(self):
        self.wb = Workbook(write_only=True)
        self.sheets = {}
        self.counts = {}
        self.skipped = {}

    def ensure_sheet(self, lvl: int):
        if lvl not in self.sheets:
            ws = self.wb.create_sheet(title=f"Level {lvl}")
            ws.append(["URL"])
            self.sheets[lvl] = ws
            self.counts[lvl] = 0
            self.skipped[lvl] = 0

    def append(self, lvl: int, url: str):
        self.ensure_sheet(lvl)
        if self.counts[lvl] >= EXCEL_MAX_DATA_ROWS:
            self.skipped[lvl] += 1
            return
        self.sheets[lvl].append([url])
        self.counts[lvl] += 1

    def save(self, out_path: Path):
        # write_only workbook needs at least one sheet
        if not self.sheets:
            ws = self.wb.create_sheet(title="Level 1")
            ws.append(["URL"])
        self.wb.save(out_path)


def pick_url_index(header_row):
    if not header_row:
        return 0
    lowered = [str(c).strip().lower() for c in header_row]
    if "url" in lowered:
        return lowered.index("url")
    return 0


def sort_csv_by_nesting(csv_path: Path, out_path: Path) -> None:
    dialect, has_header = detect_csv_dialect(csv_path)
    wb = LevelWorkbook()

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, dialect)
        url_idx = 0
        if has_header:
            header = next(reader, None)
            url_idx = pick_url_index(header)

        for row in reader:
            if not row or url_idx >= len(row):
                continue
            url = (row[url_idx] or "").strip()
            if not url or url.lower() == "nan":
                continue
            wb.append(nesting_level(url), url)

    wb.save(out_path)

    total = sum(wb.counts.values())
    print(f"Done: {out_path}")
    print(f"URLs written: {total}")
    for lvl in sorted(wb.counts):
        print(f"  Level {lvl}: {wb.counts[lvl]}")

    over = {lvl: n for lvl, n in wb.skipped.items() if n}
    if over:
        for lvl, n in sorted(over.items()):
            print(
                f"Warning: Level {lvl}: skipped {n} rows "
                f"(Excel sheet limit {EXCEL_MAX_ROWS})."
            )
        print("Extra rows were not moved to additional sheets.")


def prompt_csv_path() -> Path:
    while True:
        raw = input("CSV file path: ").strip().strip('"')
        if not raw:
            print("Please enter a path.")
            continue
        path = Path(raw).expanduser().resolve()
        if not path.exists():
            print(f"File not found: {path}")
            continue
        if path.suffix.lower() != ".csv":
            print("Please provide a .csv file.")
            continue
        return path


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Split URLs from a CSV into Excel sheets by nesting depth."
    )
    parser.add_argument(
        "csv",
        nargs="?",
        help="Input CSV (first column or URL column). If omitted, uses the first *.csv next to the script or asks interactively.",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Output Excel path (default: timestamped file next to the CSV)",
    )
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    if args.csv:
        csv_path = Path(args.csv).expanduser().resolve()
    else:
        here = Path(__file__).resolve().parent
        csv_files = sorted(here.glob("*.csv"))
        if csv_files:
            csv_path = csv_files[0]
            print(f"Using CSV next to script: {csv_path.name}")
        else:
            print("No CSV found next to the script.")
            csv_path = prompt_csv_path()

    if not csv_path.exists():
        print(f"CSV not found: {csv_path}")
        return 1

    if args.output:
        out_path = Path(args.output).expanduser().resolve()
    else:
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
        out_path = csv_path.parent / f"{ts}_sorted_urls_by_nesting.xlsx"

    try:
        sort_csv_by_nesting(csv_path, out_path)
    except Exception as e:
        print(f"Error: {e}")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
